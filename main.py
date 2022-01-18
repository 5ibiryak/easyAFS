# -*- coding: utf8 -*-
import sys
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow
import json
import os
from design import Ui_Main
import openpyxl
import atexit
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from openpyxl.styles.colors import Color

excel_created = False
num = 0
json_path = ''

# очистка json перед закрытием

def clean_json():
    pasport = {'pasport_ishodnie_dannye' : {
                        "folder_with_logs" : '',
                        "operator" : '',
                        "object_name" : '',
                        "path_for_document" : '',
                        "file_name" : '',
                        }}
    with open("data.json","w") as write_file: 
        json.dump(pasport,write_file)
    print('wow')
atexit.register(clean_json)

# работа с excel
def file_excel():
    with open ('data.json','r') as file:
        data = json.load(file)
    file_name = data['pasport_ishodnie_dannye']['file_name']
    file_path = data['pasport_ishodnie_dannye']['path_for_document']
    os.chdir(file_path)

# основной класс

class Main(QMainWindow, Ui_Main):
    
    
    def __init__(self, parent=None):
        super(Main, self).__init__(parent)
        self.setupUi(self)
        
        self.pushButton1.clicked.connect(self.ishodnie_dannye)
        self.pushButton2.clicked.connect(self.pasport_AFS_1)
        self.pushButton3.clicked.connect(self.jurnal_AFS)
        


    def main_menu(self):
        self.QtStack.setCurrentIndex(0)
        



########                              ############
########                              ############
#       страница с исходными данными
########                              ############
########                              ############

    def ishodnie_dannye(self):
        self.QtStack.setCurrentIndex(1)
        
        def getDirectory_logs():
            dirlist = QtWidgets.QFileDialog.getExistingDirectory()
            self.plainTextEdit_folder.setPlainText(format(dirlist))

        def getDirectory_path_document():
            dirlist = QtWidgets.QFileDialog.getExistingDirectory()
            self.plainTextEdit_folder_4.setPlainText(format(dirlist))

        def btn_click():
            # СЧИТЫВАНИЕ ДАННЫХ
            
            folder_with_logs = self.plainTextEdit_folder.toPlainText() # папка с логами
            operator = self.plainTextEdit_folder_2.toPlainText() # ФИО оператора 
            object_name = self.plainTextEdit_folder_3.toPlainText() # наименование объекта
            path_for_document = self.plainTextEdit_folder_4.toPlainText() # путь до создоваемого файла
            file_name = self.plainTextEdit_folder_5.toPlainText() # название excel файла
            
            pasport = {'pasport_ishodnie_dannye' : {
            "folder_with_logs" : folder_with_logs,
            "operator" : operator,
            "object_name" : object_name,
            "path_for_document" : path_for_document,
            "file_name" : file_name,
            }}
            global json_path
            json_path = os.getcwd()
            print(json_path)
            with open("data.json","w") as write_file: 
                json.dump(pasport,write_file) 
            self.main_menu()
            


        if os.path.exists('data.json')==True:
            try:   
                with open('data.json') as file:
                    file_content = file.read().strip()
            # Проверяем, пустой ли файл
                if not file_content:
                    pasport = {'pasport_ishodnie_dannye' : {
                        "folder_with_logs" : '',
                        "operator" : '',
                        "object_name" : '',
                        "path_for_document" : '',
                        "file_name" : '',
                        }}
                    with open("data.json","w") as write_file: 
                        json.dump(pasport,write_file) 
            
            
            except FileNotFoundError:
                pass


            
        else:
            pasport = {'pasport_ishodnie_dannye' : {
                        "folder_with_logs" : '',
                        "operator" : '',
                        "object_name" : '',
                        "path_for_document" : '',
                        "file_name" : '',
                        }}
            with open("data.json","w") as write_file: 
                json.dump(pasport,write_file) 

        try:
            with open("data.json", "r") as read_file:
                data = json.load(read_file)
        except:
            pass


        try:
            self.plainTextEdit_folder.setPlainText(data["pasport_ishodnie_dannye"]["folder_with_logs"])
            self.plainTextEdit_folder_2.setPlainText(data["pasport_ishodnie_dannye"]["operator"])
            self.plainTextEdit_folder_3.setPlainText(data["pasport_ishodnie_dannye"]["object_name"])
            self.plainTextEdit_folder_4.setPlainText(data["pasport_ishodnie_dannye"]["path_for_document"])
            self.plainTextEdit_folder_5.setPlainText(data["pasport_ishodnie_dannye"]["file_name"])
        except:
            self.plainTextEdit_folder.setPlainText("")
            self.plainTextEdit_folder_2.setPlainText("")
            self.plainTextEdit_folder_3.setPlainText("")
            self.plainTextEdit_folder_4.setPlainText("")
            self.plainTextEdit_folder_5.setPlainText("")

        self.pushButton_folder.clicked.connect(getDirectory_logs)
        self.pushButton_folder_2.clicked.connect(getDirectory_path_document)
        self.pushButton_download.clicked.connect(btn_click)



########                              ############
########                              ############
#         страница с паспортом 1
########                              ############
########                              ############

#сделать проверку есть ли уже такой афс или нет
    def pasport_AFS_1(self):
        self.QtStack.setCurrentIndex(2)
        self.pushButton.clicked.connect(self.main_menu)

        def btn_next():
            global Mission_number
            Mission_number = self.comboBox.currentText()
            Date = self.plainTextEdit_data.toPlainText()
            Time = self.plainTextEdit_time.toPlainText()
            AFS_type = self.comboBox_2.currentText()
            AFS_mode = self.comboBox_3.currentText()
            UMA_name = self.comboBox_4.currentText()
            page_one = {'AFS_' + Mission_number:{
                "Mission_number":Mission_number,
                "Date":Date,
                "Time":Time,
                "AFS_type":AFS_type,
                "AFS_mode":AFS_mode,
                "UMA_name":UMA_name,
            }}
            ### директория json
            global json_path
            os.chdir(json_path)
            with open ('data.json','r') as file:
                data = json.load(file)
            data.update(page_one)
            with open('data.json',"w") as file:
                json.dump(data,file)
            self.pasport_AFS_2()


        self.pushButton_2.clicked.connect(btn_next)


########                              ############
########                              ############
#         страница с паспортом 2
########                              ############
########                              ############


    def pasport_AFS_2(self):
        self.QtStack.setCurrentIndex(3)
        self.pushButton_AFS_2.clicked.connect(self.pasport_AFS_1)

 
        def btn_next():
            registry_number = self.plainTextEdit_AFS_2.toPlainText()
            pay_load_1 = self.plainTextEdit_AFS_3.toPlainText()
            pay_load_2 = self.plainTextEdit_AFS_4.toPlainText()
            mission_software = self.plainTextEdit_AFS_5.toPlainText()
            solution_method = self.comboBox_AFS_2_1.currentText()
            altitude = self.plainTextEdit_AFS_6.toPlainText()
            page_two = {
                "registry_number":registry_number,
                "pay_load_1":pay_load_1,
                "pay_load_2":pay_load_2,
                "mission_software":mission_software,
                "solution_method":solution_method,
                "altitude":altitude,
            }
            ### директория json
            global json_path
            os.chdir(json_path)
            with open ('data.json','r') as file:
                data = json.load(file)
            data['AFS_' + Mission_number].update(page_two)
            with open('data.json',"w") as file:
                json.dump(data,file)
            self.pasport_AFS_3()

        self.pushButton_AFS_2_2.clicked.connect(btn_next)
        


########                              ############
########                              ############
#         страница с паспортом 3
########                              ############
########                              ############


    def pasport_AFS_3(self):
        self.QtStack.setCurrentIndex(4)
        self.pushButton_AFS_3_1.clicked.connect(self.pasport_AFS_2)


        def btn_next():
            horizontal_lap  = self.plainTextEdit_AFS_3_1.toPlainText()
            vertical_lap  = self.plainTextEdit_AFS_3_3.toPlainText()
            shape = self.plainTextEdit_AFS_3_2.toPlainText()
            shots_number = self.plainTextEdit_AFS_3_4.toPlainText()
            precipitation = self.comboBox_AFS_3_1.currentText()
            undercast = self.comboBox_AFS_3_2.currentText()
            page_three = {
                "horizontal_lap":horizontal_lap,
                "vertical_lap":vertical_lap,
                "shape":shape,
                "shots_number":shots_number,
                "precipitation":precipitation,
                "undercast":undercast,
            }
            ### директория json
            global json_path
            os.chdir(json_path)
            with open ('data.json','r') as file:
                data = json.load(file)
            data['AFS_' + Mission_number].update(page_three)
            with open('data.json',"w") as file:
                json.dump(data,file)
            self.geodeziy()


        self.pushButton_AFS_3_2.clicked.connect(btn_next)


########                              ############
########                              ############
#         страница с геодезией
########                              ############
########                              ############

    def geodeziy(self):
        self.QtStack.setCurrentIndex(5)
        self.pushButton_geodeziy_1.clicked.connect(self.pasport_AFS_3)

        def btn_next():
            home_point  = self.plainTextEdit_geodeziy_1.toPlainText()
            device   = self.plainTextEdit_geodeziy_3.toPlainText()
            log_number = self.plainTextEdit_geodeziy_2.toPlainText()
            device_high = self.plainTextEdit_geodeziy_4.toPlainText()
            file_name  = self.plainTextEdit_geodeziy_5.toPlainText()
            
            page_four = {
                "home_point":home_point,
                "device":device,
                "log_number":log_number,
                "device_high":device_high,
                "file_name":file_name,
            }
            ### директория json
            global json_path
            os.chdir(json_path)
            with open ('data.json','r') as file:
                data = json.load(file)
            data['AFS_' + Mission_number].update(page_four)
            with open('data.json',"w") as file:
                json.dump(data,file)
            self.primechania()

        self.pushButton_geodeziy_2.clicked.connect(btn_next)

########                              ############
########                              ############
#         страница с примечаниями
########                              ############
########                              ############

    def primechania(self):
        self.QtStack.setCurrentIndex(6)
        self.pushButton_primechania_1.clicked.connect(self.geodeziy)
        global excel_created
        def btn_next():
            processing_usage  = self.comboBox_primechania.currentText()
            usage_problem   = self.plainTextEdit_primechania_2.toPlainText()
            incidents = self.plainTextEdit_primechania_1.toPlainText()

            
            page_five = {
                "processing_usage":processing_usage,
                "usage_problem":usage_problem,
                "incidents":incidents,
            }
            ### директория json
            global json_path
            os.chdir(json_path)
            with open ('data.json','r') as file:
                data = json.load(file)
            data['AFS_' + Mission_number].update(page_five)
            with open('data.json',"w") as file:
                json.dump(data,file)
            global excel_created
            if (excel_created!=True):
                create_excel_doc()
                add_styles_to_wb()
                add_journal()
                excel_created = True
            num = 0
            #print(Mission_number)
            add_ws(Mission_number)
            
            self.main_menu()
            
        self.pushButton_primechania_2.clicked.connect(btn_next)
        
        # создание таблицы excel
        def create_excel_doc():
            file_name = ''
            file_path = ''
            with open ('data.json','r') as file:
                data = json.load(file)
                file_name = data["pasport_ishodnie_dannye"]["file_name"]
                file_path = data["pasport_ishodnie_dannye"]["path_for_document"]
            file_name = file_name + ".xlsx"
            if os.path.exists(file_path+"/"+file_name):
                # файл существует
                #print("Файл существует")
                pass
            else:
                # файл не существует
                os.chdir(file_path)
                workbook = xlsxwriter.Workbook(file_name)
                workbook.close()
        
        # добавление стилей в воркбук
        def add_styles_to_wb():
            file_name = ''
            file_path = ''
            ### директория json
            global json_path
            os.chdir(json_path)
            with open ('data.json','r') as file:
                data = json.load(file)
                file_name = data["pasport_ishodnie_dannye"]["file_name"]
                file_path = data["pasport_ishodnie_dannye"]["path_for_document"]
            file_name = file_name + ".xlsx"
            os.chdir(file_path)
            try:
                myfile = open(file_name, "r+")
            except IOError:
                print("Файл открыт")
            else:
                workbook = load_workbook(file_name)
                
                highlight_blue = NamedStyle(name="highlight_blue")
                highlight_blue.font = Font(name = 'Arial', bold=True, size=11)
                blueFill = PatternFill(start_color='0099CCFF',
                   end_color='0099CCFF',
                   fill_type='solid')
                highlight_blue.fill = blueFill
                workbook.add_named_style(highlight_blue)
                
                highlight_gray = NamedStyle(name="highlight_gray")
                highlight_gray.font = Font(name = 'Arial', bold=True, size=11)
                grayFill = PatternFill(start_color='00DCDCDC',
                   end_color='00DCDCDC',
                   fill_type='solid')
                highlight_gray.fill = grayFill
                workbook.add_named_style(highlight_gray)
                
                highlight_gray2 = NamedStyle(name="highlight_gray2")
                highlight_gray2.font = Font(name = 'Arial', bold=False, size=11)
                grayFill2 = PatternFill(start_color='00DCDCDC',
                   end_color='00DCDCDC',
                   fill_type='solid')
                highlight_gray2.fill = grayFill2
                workbook.add_named_style(highlight_gray2)
                
                usual_style = NamedStyle(name="usual_style")
                usual_style.font = Font(name = 'Arial', bold=False, size=11)
                usual_style.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True, shrink_to_fit=False)
                workbook.add_named_style(usual_style)
                
                jtitle_style = NamedStyle(name="jtitle_style")
                jtitle_style.font = Font(name = 'Arial', bold=True, size=11)
                jtitle_style.alignment = Alignment(horizontal='center',vertical='center', wrap_text=True, shrink_to_fit=False)
                jtitle_style.border = Border(left=Side(border_style='medium', color='FF000000'), right=Side(border_style='medium', color='FF000000'), top=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='medium',color='FF000000'))
                workbook.add_named_style(jtitle_style)
                
                rotated_style = NamedStyle(name="rotated_style")
                rotated_style.font = Font(name = 'Arial', bold=True, size=11)
                rotated_style.alignment = Alignment(horizontal='center',vertical='center', wrap_text=True, shrink_to_fit=False, textRotation=90)
                rotated_style.border = Border(left=Side(border_style='medium', color='FF000000'), right=Side(border_style='medium', color='FF000000'), top=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='medium',color='FF000000'))
                workbook.add_named_style(rotated_style)
                
                # сохранить и закрыть файл
                workbook.save(filename=file_name)
                workbook.close()
                
        #добавление журнала полетов
        def add_journal():
            file_name = ''
            file_path = ''
            ### директория json
            global json_path
            os.chdir(json_path)
            with open ('data.json','r') as file:
                data = json.load(file)
                file_name = data["pasport_ishodnie_dannye"]["file_name"]
                file_path = data["pasport_ishodnie_dannye"]["path_for_document"]
            file_name = file_name + ".xlsx"
            os.chdir(file_path)
            try:
                myfile = open(file_name, "r+")
            except IOError:
                print("Файл открыт")
            else:
                workbook = load_workbook(file_name)
                sheet_name = "Журнал полетов"
                #print(sheet_name)
                if sheet_name in workbook.sheetnames:
                    pass
                else:
                    worksheet_journal = workbook.create_sheet(sheet_name)
                    row = 1
                    column = 2
                    frame_1 = ["Дата полета", "БВС", "№ борта", "№ полета", "Наименование объекта", "Тип съемки", "Вид съемки", "S/n камеры"]
                    frame_2 = ["левая/ надирная", "правая"]
                    frame_3 = ["Для обработки", "Примечания", "Полетная база 1"]
                    frame_4 = ["порядковый номер точки", "прибор (назв, номер)", "порядковый номер лога", "высота прибора, мм, вертикальная, наклонная", "название файла"]
                    frame_5 = ["Время полета", "Количество фотографий"]
                    
                    for item in frame_1 :
                        worksheet_journal.cell(row=row, column=column).value = item
                        column += 1
                    
                    worksheet_journal.merge_cells('A1:A3')
                    worksheet_journal.column_dimensions['A'].width = float(8.09+0.71)
                    worksheet_journal.merge_cells('B1:B3')
                    worksheet_journal.column_dimensions['B'].width = float(12.55+0.71)
                    worksheet_journal.merge_cells('C1:C3')
                    worksheet_journal.column_dimensions['C'].width = float(12.55+0.71)
                    worksheet_journal.merge_cells('D1:D3')
                    worksheet_journal.column_dimensions['D'].width = float(12.55+0.71)
                    worksheet_journal.merge_cells('E1:E3')
                    worksheet_journal.column_dimensions['E'].width = float(12.55+0.71)
                    worksheet_journal.merge_cells('F1:F3')
                    worksheet_journal.column_dimensions['F'].width = float(42.27+0.71)
                    worksheet_journal.merge_cells('G1:G3')
                    worksheet_journal.column_dimensions['G'].width = float(12.55+0.71)
                    worksheet_journal.merge_cells('H1:H3')
                    worksheet_journal.column_dimensions['H'].width = float(8.09+0.71)
                    worksheet_journal.merge_cells('I1:J1')
                    worksheet_journal.column_dimensions['I'].width = float(23.64+0.71)
                    worksheet_journal.column_dimensions['J'].width = float(12.91+0.71)
                    
                    row = 2
                    column = 9
                    for item in frame_2 :
                        worksheet_journal.cell(row=row, column=column).value = item
                        column += 1
                    
                    worksheet_journal.merge_cells('I2:I3')
                    worksheet_journal.merge_cells('J2:J3')
                    
                    row = 1
                    column = 11
                    for item in frame_3 :
                        worksheet_journal.cell(row=row, column=column).value = item
                        column += 1
                    
                    worksheet_journal.merge_cells('K1:K3')
                    worksheet_journal.column_dimensions['K'].width = float(17.55+0.71)
                    worksheet_journal.merge_cells('L1:L3')
                    worksheet_journal.column_dimensions['L'].width = float(17.55+0.71)
                    worksheet_journal.merge_cells('M1:Q1')
                    worksheet_journal.column_dimensions['M'].width = float(10.91+0.71)
                    worksheet_journal.column_dimensions['N'].width = float(10.91+0.71)
                    worksheet_journal.column_dimensions['O'].width = float(14.18+0.71)
                    worksheet_journal.column_dimensions['P'].width = float(17.91+0.71)
                    worksheet_journal.column_dimensions['Q'].width = float(21.18+0.71)
                    worksheet_journal.merge_cells('N2:P2')
                    worksheet_journal.merge_cells('M2:M3')
                    worksheet_journal.merge_cells('Q2:Q3')
                    
                    worksheet_journal['M2'].value = frame_4[0]
                    worksheet_journal['N3'].value = frame_4[1]
                    worksheet_journal['O3'].value = frame_4[2]
                    worksheet_journal['P3'].value = frame_4[3]
                    worksheet_journal['Q2'].value = frame_4[4]
                    
                    row = 1
                    column = 18
                    for item in frame_5 :
                        worksheet_journal.cell(row=row, column=column).value = item
                        column += 1
                    
                    worksheet_journal.merge_cells('R1:R3')
                    worksheet_journal.column_dimensions['R'].width = float(12.36+0.71)
                    worksheet_journal.merge_cells('S1:S3')
                    worksheet_journal.column_dimensions['S'].width = float(8.09+0.71)
                    
                    for c in range(2, 20) :
                        for r in range(1, 4):
                            worksheet_journal.cell(row=r, column=c).style = 'jtitle_style'
                    
                    for c in range(2, 6) :
                        for r in range(1, 2):
                            worksheet_journal.cell(row=r, column=c).style = 'rotated_style'
                    worksheet_journal['M2'].style = 'rotated_style'
                    worksheet_journal['R1'].style = 'rotated_style'
                    worksheet_journal['S1'].style = 'rotated_style'
                    worksheet_journal.row_dimensions[3].height = float(62.0+0.71)
                    
                    workbook.save(filename=file_name)
                    workbook.close()
                
        # добавление листа в таблицу (отдельный АФС)
        def add_ws(num):
            previous_path=os.getcwd()
            num = num
            file_name = ''
            file_path = ''
            ### директория json
            global json_path
            os.chdir(json_path)
            with open ('data.json','r') as file:
                data = json.load(file)
                file_name = data["pasport_ishodnie_dannye"]["file_name"]
                file_path = data["pasport_ishodnie_dannye"]["path_for_document"]
            file_name = file_name + ".xlsx"
            os.chdir(file_path)
            try:
                myfile = open(file_name, "r+")
            except IOError:
                print("Файл открыт")
            else:
                workbook = load_workbook(file_name)
                sheet_name = "АФС_"+str(num)
                #print(sheet_name)
                if sheet_name in workbook.sheetnames:
                    pass
                else:
                    worksheet_AFS1 = workbook.create_sheet(sheet_name)
                    row = 1
                    column = 1
                    frame_1 = ["Наименование объекта", "Оператор", "Номер полета", "Дата полета", "Время полета", "Тип АФС", "Вид АФС", "Название БВС", "Регистрационный номер борта", "Полезная нагрузка 1", "Полезная нагрузка 2","ПО для планирования полета", "Метод решения", "Высота полета", "Продольное перекрытие", "Поперечное перекрытие", "Разрешение", "Количество снимков", "Осадки", "Облачность"]
                    frame_2 = ["Геодезия", "Наименование точки (базы)", "Прибор (название, номер)", "Порядковый номер лога (базы)", "Высота прибора (мм)", "Название файла"]
                    frame_3 = ["Примечания", "Использование полета в обработке", "Причина, по которой нельзя использовать", "Происшествия"]
                    
                    for item in frame_1 :
                        worksheet_AFS1.cell(row=row, column=column).value = item
                        row += 1
                    
                    row = 22
                    column = 1
                    for item in frame_2 :
                        worksheet_AFS1.cell(row=row, column=column).value = item
                        row += 1
                    
                    row = 29
                    column = 1
                    for item in frame_3 :
                        worksheet_AFS1.cell(row=row, column=column).value = item
                        row += 1
                    
                    worksheet_AFS1.merge_cells('A22:B22')
                    worksheet_AFS1.merge_cells('A29:B29')
                    
                    #setting width of columns
                    worksheet_AFS1.column_dimensions['A'].width = float(48.6)
                    worksheet_AFS1.column_dimensions['B'].width = float(56.9)
                    
                    
                    
                    for c in range(1, 3) :
                        for r in range(1, 33):
                            worksheet_AFS1.cell(row=r, column=c).style = 'usual_style'
                    
                    
                    worksheet_AFS1['A22'].style = 'highlight_blue'
                    worksheet_AFS1['A29'].style = 'highlight_blue'
                    worksheet_AFS1['A1'].style = 'highlight_gray'
                    worksheet_AFS1['A2'].style = 'highlight_gray'
                    worksheet_AFS1['B1'].style = 'highlight_gray'
                    worksheet_AFS1['B2'].style = 'highlight_gray'
                    
                    bd = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin',color='FF000000'))
                    
                    for c in range(1, 3) :
                        for r in range(1, 21):
                            worksheet_AFS1.cell(row=r, column=c).border = bd
                            
                    for c in range(1, 3) :
                        for r in range(22, 28):
                            worksheet_AFS1.cell(row=r, column=c).border = bd
                                
                    for c in range(1, 3) :
                        for r in range(29, 33):
                            worksheet_AFS1.cell(row=r, column=c).border = bd
                    
                    fontItalic = Font(name='Arial', size=11, bold=False, italic=True, color='00909090')
                    for c in range(2, 3) :
                        for r in range(1, 21):
                            worksheet_AFS1.cell(row=r, column=c).font = fontItalic
                    for c in range(2, 3) :
                        for r in range(23, 28):
                            worksheet_AFS1.cell(row=r, column=c).font = fontItalic
                    for c in range(2, 3) :
                        for r in range(30, 33):
                            worksheet_AFS1.cell(row=r, column=c).font = fontItalic
                    
                    # добавление на лист данных из json
                    AFS_name = "AFS_"+str(num)
                    worksheet_AFS1['B1'].value = data["pasport_ishodnie_dannye"]["object_name"]
                    worksheet_AFS1['B2'].value = data["pasport_ishodnie_dannye"]["operator"]
                    worksheet_AFS1['B3'].value = data[AFS_name]["Mission_number"]
                    worksheet_AFS1['B4'].value = data[AFS_name]["Date"]
                    worksheet_AFS1['B5'].value = data[AFS_name]["Time"]
                    worksheet_AFS1['B6'].value = data[AFS_name]["AFS_type"]
                    worksheet_AFS1['B7'].value = data[AFS_name]["AFS_mode"]
                    worksheet_AFS1['B8'].value = data[AFS_name]["UMA_name"]
                    worksheet_AFS1['B9'].value = data[AFS_name]["registry_number"]
                    worksheet_AFS1['B10'].value = data[AFS_name]["pay_load_1"]
                    worksheet_AFS1['B11'].value = data[AFS_name]["pay_load_2"]
                    worksheet_AFS1['B12'].value = data[AFS_name]["mission_software"]
                    worksheet_AFS1['B13'].value = data[AFS_name]["solution_method"]
                    worksheet_AFS1['B14'].value = data[AFS_name]["altitude"]
                    worksheet_AFS1['B15'].value = data[AFS_name]["horizontal_lap"]
                    worksheet_AFS1['B16'].value = data[AFS_name]["vertical_lap"]
                    worksheet_AFS1['B17'].value = data[AFS_name]["shape"]
                    worksheet_AFS1['B18'].value = data[AFS_name]["shots_number"]
                    worksheet_AFS1['B19'].value = data[AFS_name]["precipitation"]
                    worksheet_AFS1['B20'].value = data[AFS_name]["undercast"]
                    
                    worksheet_AFS1['B23'].value = data[AFS_name]["home_point"]
                    worksheet_AFS1['B24'].value = data[AFS_name]["device"]
                    worksheet_AFS1['B25'].value = data[AFS_name]["log_number"]
                    worksheet_AFS1['B26'].value = data[AFS_name]["device_high"]
                    worksheet_AFS1['B27'].value = data[AFS_name]["file_name"]
                    
                    worksheet_AFS1['B30'].value = data[AFS_name]["processing_usage"]
                    worksheet_AFS1['B31'].value = data[AFS_name]["usage_problem"]
                    worksheet_AFS1['B32'].value = data[AFS_name]["incidents"]
                    
                    ###добавление записи об АФС в журнал полетов
                    worksheet_journal = workbook["Журнал полетов"]
                    cur_row = int(num)+3
                    for c in range(1, 20) :
                        for r in range(cur_row, cur_row+1):
                            worksheet_journal.cell(row=r, column=c).style = 'usual_style'
                            worksheet_journal.cell(row=r, column=c).border = bd
                            
                    worksheet_journal['A'+str(cur_row)].value = 'АФС_'+str(num)
                    worksheet_journal['B'+str(cur_row)].value = data[AFS_name]["Date"]
                    worksheet_journal['C'+str(cur_row)].value = data[AFS_name]["UMA_name"]
                    worksheet_journal['D'+str(cur_row)].value = data[AFS_name]["registry_number"]
                    worksheet_journal['E'+str(cur_row)].value = data[AFS_name]["Mission_number"]
                    worksheet_journal['F'+str(cur_row)].value = data["pasport_ishodnie_dannye"]["object_name"]
                    worksheet_journal['G'+str(cur_row)].value = data[AFS_name]["AFS_type"]
                    worksheet_journal['H'+str(cur_row)].value = data[AFS_name]["AFS_mode"]
                    worksheet_journal['I'+str(cur_row)].value = data[AFS_name]["pay_load_1"]
                    worksheet_journal['J'+str(cur_row)].value = data[AFS_name]["pay_load_2"]
                    worksheet_journal['K'+str(cur_row)].value = data[AFS_name]["processing_usage"]
                    worksheet_journal['L'+str(cur_row)].value = ''
                    worksheet_journal['M'+str(cur_row)].value = data[AFS_name]["home_point"]
                    worksheet_journal['N'+str(cur_row)].value = data[AFS_name]["device"]
                    worksheet_journal['O'+str(cur_row)].value = data[AFS_name]["log_number"]
                    worksheet_journal['P'+str(cur_row)].value = data[AFS_name]["device_high"]
                    worksheet_journal['Q'+str(cur_row)].value = data[AFS_name]["file_name"]
                    worksheet_journal['R'+str(cur_row)].value = data[AFS_name]["Time"]
                    worksheet_journal['S'+str(cur_row)].value = data[AFS_name]["shots_number"]
                    
                    if 'Sheet1' in workbook.sheetnames:
                        workbook.remove(workbook['Sheet1'])
                    
                    
                    workbook.save(filename=file_name)
                    workbook.close()
            os.chdir(previous_path)


########                              ############
########                              ############
#         страница с геодезией
########                              ############
########                              ############

    def jurnal_AFS(self):
        self.QtStack.setCurrentIndex(7)
        self.pushButton_jurnal.clicked.connect(self.main_menu)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    showMain = Main()
    sys.exit(app.exec_())











